# Для решения этой задачи необходимо установить библиотеку openpyxl и создать скрипт со следующем содержанием:
# Запустите скрипт и в качестве ответа введите то, что он выведет.

import openpyxl

wb = openpyxl.load_workbook('inputs/input12.xlsx')
sh = wb.active
nmin = sh.cell(row=7, column=2).value
for rownum in range(8, 28):
    nmin = min(nmin, sh.cell(row=rownum, column=2).value)
print(nmin)
